"""
Excel export service using openpyxl.
Creates styled .xlsx workbooks with auto-filters and column widths.
"""
from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
DATE_FMT = "DD.MM.YYYY HH:MM"
MONEY_FMT = '#,##0.00" ₽"'
QTY_FMT = "#,##0.###"


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = str(cell.value or "")
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)


def _style_header(ws, col_count: int) -> None:
    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"


def build_requests_xlsx(rows: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"
    headers = ["№ заявки", "Дата", "Подразделение", "Товаровед", "Позиция",
               "Ед.", "Кол-во", "Цена ед.", "Сумма", "Статус", "Комментарий"]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("request_id"), r.get("created_at"), r.get("branch"),
            r.get("created_by"), r.get("item"), r.get("unit"),
            r.get("qty_requested"), r.get("unit_cost"), r.get("requested_cost"),
            r.get("status"), r.get("comment", ""),
        ])
    _style_header(ws, len(headers))
    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_deliveries_xlsx(rows: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Доставки"
    headers = ["№ визита", "Начало", "Окончание", "Водитель", "Подразделение",
               "Позиция", "Ед.", "Было", "Доставлено", "Сумма", "Осталось",
               "Результат", "Причина недовоза", "Статус"]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("session_id"), r.get("started_at"), r.get("finished_at"),
            r.get("driver"), r.get("branch"), r.get("item"), r.get("unit"),
            r.get("qty_before"), r.get("qty_delivered_now"), r.get("delivered_cost"),
            r.get("qty_after"), r.get("result_status"), r.get("shortage_reason", ""),
            r.get("session_status"),
        ])
    _style_header(ws, len(headers))
    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_demand_xlsx(rows: list[dict]) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Потребности"
    headers = ["Подразделение", "Адрес", "Позиция", "Ед.", "Цена ед.",
               "Запрошено", "Доставлено", "Остаток", "Сумма остатка", "Статус"]
    ws.append(headers)
    for r in rows:
        ws.append([
            r.get("branch"), r.get("address"), r.get("item"), r.get("unit"),
            r.get("unit_cost"), r.get("qty_requested"), r.get("qty_delivered"),
            r.get("qty_remaining"), r.get("remaining_cost"), r.get("status"),
        ])
    _style_header(ws, len(headers))
    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_summary_xlsx(analytics: dict, date_from: str, date_to: str) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка"
    ws.append(["Период", f"{date_from} — {date_to}"])
    ws.append([])
    ws.append(["Показатель", "Значение"])
    metrics = [
        ("Филиалов с потребностью", analytics.get("active_branches", 0)),
        ("Активных позиций", analytics.get("active_positions", 0)),
        ("Остаток к доставке (ед.)", analytics.get("active_qty", 0)),
        ("Оценка потребности (₽)", analytics.get("active_cost", 0)),
        ("Запрошено всего (ед.)", analytics.get("requested_qty", 0)),
        ("Запрошено всего (₽)", analytics.get("requested_cost", 0)),
        ("Доставлено всего (ед.)", analytics.get("delivered_qty", 0)),
        ("Доставлено всего (₽)", analytics.get("delivered_cost", 0)),
        ("Визитов водителей", analytics.get("delivery_sessions_count", 0)),
    ]
    for label, value in metrics:
        ws.append([label, round(float(value), 2)])
    for col_idx in range(1, 3):
        cell = ws.cell(row=3, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)
    _auto_width(ws)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
