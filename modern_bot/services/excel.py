import asyncio
from typing import List, Any, Dict
from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime
from modern_bot.config import EXCEL_FILE, EXCEL_HEADERS, DOCS_DIR
from modern_bot.utils.files import sanitize_filename
import logging

logger = logging.getLogger(__name__)
excel_lock = asyncio.Lock()

async def read_excel_data() -> List[List[str]]:
    """Reads data from Excel file safely."""
    def _read_excel():
        if not EXCEL_FILE.exists():
            return []
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
        wb.close()
        return rows
    
    async with excel_lock:
        return await asyncio.to_thread(_read_excel)

async def update_excel(data: Dict[str, Any]) -> None:
    """Updates Excel file with new conclusion data."""
    def _write_excel():
        if not EXCEL_FILE.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(EXCEL_HEADERS)
        else:
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active

        items = data.get("photo_desc", [])
        for idx, item in enumerate(items, 1):
            row = [
                data.get("ticket_number", "Не указано"),
                data.get("issue_number", "Не указано"),
                data.get("department_number", "Не указано"),
                data.get("date", "Не указано"),
                data.get("region", "Не указано"),
                idx,
                item.get("description", "Нет описания"),
                item.get("evaluation", "Нет данных"),
                data.get("user_name", "Unknown")
            ]
            ws.append(row)
        wb.save(EXCEL_FILE)
        wb.close()

    async with excel_lock:
        await asyncio.to_thread(_write_excel)
        logger.info("Excel file updated.")

async def create_excel_snapshot(rows: List[List[Any]], filename_prefix: str) -> Path:
    """Creates a temporary Excel snapshot."""
    DOCS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%d.%m.%Y_%H-%M-%S")

    def _write_snapshot() -> Path:
        wb = Workbook()
        ws = wb.active
        ws.append(EXCEL_HEADERS)
        for row in rows:
            ws.append(row)
        raw_name = f"{filename_prefix}_{timestamp}.xlsx"
        filepath = DOCS_DIR / sanitize_filename(raw_name)
        wb.save(filepath)
        wb.close()
        return filepath

    return await asyncio.to_thread(_write_snapshot)

async def prune_excel_data(cutoff: datetime) -> int:
    """Remove rows older than cutoff and rewrite Excel."""
    def _prune() -> int:
        if not EXCEL_FILE.exists():
            return 0

        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        kept: List[List[Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 4:
                continue
            date_val = row[3]
            dt = None
            if isinstance(date_val, datetime):
                dt = date_val
            else:
                try:
                    dt = datetime.strptime(str(date_val), "%d.%m.%Y")
                except (ValueError, TypeError):
                    dt = None

            if dt and dt >= cutoff:
                kept.append(list(row))
        wb.close()

        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.append(EXCEL_HEADERS)
        for row in kept:
            new_ws.append(row)
        new_wb.save(EXCEL_FILE)
        new_wb.close()
        return len(kept)

    async with excel_lock:
        return await asyncio.to_thread(_prune)
