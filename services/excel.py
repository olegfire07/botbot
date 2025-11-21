import asyncio
from typing import List, Any, Optional
from datetime import datetime
from openpyxl import Workbook, load_workbook
from config.settings import settings
from utils.helpers import parse_date_str, sanitize_filename
import logging

logger = logging.getLogger(__name__)
excel_lock = asyncio.Lock()

async def read_excel_data() -> List[List[str]]:
    """Asynchronously reads data from Excel file."""
    def _read_excel():
        if not settings.EXCEL_FILE.exists():
            return []
        wb = load_workbook(settings.EXCEL_FILE)
        ws = wb.active
        rows = [list(row) for row in ws.iter_rows(min_row=2, values_only=True)]
        wb.close()
        return rows
    
    async with excel_lock:
        return await asyncio.to_thread(_read_excel)

async def update_excel(data: dict) -> None:
    """Updates Excel file with conclusion info."""
    def _write_excel():
        if not settings.EXCEL_FILE.exists():
            wb = Workbook()
            ws = wb.active
            ws.append(settings.EXCEL_HEADERS)
        else:
            wb = load_workbook(settings.EXCEL_FILE)
            ws = wb.active

        items = data.get("photo_desc", [])
        for idx, item in enumerate(items, 1):
            row = [
                data.get("ticket_number", "Не указано"),
                data.get("issue_number", "Не указано"),
                data.get("department_number", "Не указано"),
                data.get("date", "Не указано"),
                data.get("region", "Не указано"),
                idx,
                item.get("description", "Нет описания"),
                item.get("evaluation", "Нет данных")
            ]
            ws.append(row)
        wb.save(settings.EXCEL_FILE)
        wb.close()

    async with excel_lock:
        await asyncio.to_thread(_write_excel)
        logger.info("Excel file updated.")

async def filter_records(start_date: Optional[datetime] = None, end_date: Optional[datetime] = None, region: Optional[str] = None) -> List[List[Any]]:
    records = await read_excel_data()
    if not records:
        return []

    filtered = []
    for row in records:
        date_obj = parse_date_str(row[3])
        if start_date and (not date_obj or date_obj < start_date):
            continue
        if end_date and (not date_obj or date_obj > end_date):
            continue
        if region and (row[4] or "").strip() != region:
            continue
        filtered.append(row)
    return filtered

async def create_excel_snapshot(rows: List[List[Any]], filename_prefix: str) -> Any:
    """Creates a temp Excel file with given rows."""
    settings.DOCS_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%d.%m.%Y_%H-%M-%S")

    def _write_snapshot():
        wb = Workbook()
        ws = wb.active
        ws.append(settings.EXCEL_HEADERS)
        for row in rows:
            ws.append(row)
        raw_name = f"{filename_prefix}_{timestamp}.xlsx"
        filepath = settings.DOCS_DIR / sanitize_filename(raw_name)
        wb.save(filepath)
        wb.close()
        return filepath

    return await asyncio.to_thread(_write_snapshot)
